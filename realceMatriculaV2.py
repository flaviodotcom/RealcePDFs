import os
import fitz
import openpyxl
import customtkinter as ctk
from tkinter import filedialog

file_name = ""

# Função para selecionar o arquivo Excel
def select_excel_file():
    file_path = filedialog.askopenfilename(filetypes=[("Excel Files", "*.xlsx"), ("CSV Files", "*.csv"), ("Excel 97-2003 Files", "*.xls")])
    excel_file_entry.delete(0, ctk.END)
    excel_file_entry.insert(0, file_path)

# Função para selecionar o arquivo PDF
def select_pdf_file():
    file_path = filedialog.askopenfilename(filetypes=[("PDF Files", "*.pdf")])
    global file_name
    pdf_file_entry.delete(0, ctk.END)
    pdf_file_entry.insert(0, file_path)
    file_name = os.path.basename(file_path)

def toggle_info_panel():
    if info_frame.grid_info():
        info_frame.grid_remove()
    else:
        info_frame.grid(row=3, column=0, columnspan=3, padx=10, pady=10, sticky=ctk.EW)

# Função para destacar os números de matrícula no PDF
def highlight_registration_numbers():
    global file_name
    # abre o arquivo PDF
    pdf_file = fitz.open(pdf_file_entry.get())

    # abre a planilha do Excel
    excel_file = openpyxl.load_workbook(excel_file_entry.get())
    worksheet = excel_file.active

    # obtem o número de linhas na planilha
    num_rows = worksheet.max_row

    # percorre todas as linhas da planilha
    for row in range(1, num_rows + 1):
        # obtem o número da matrícula na coluna B da planilha
        registration_number = worksheet.cell(row=row, column=2).value
        registration_number = str(registration_number)  # converte para string

        # percorre todas as páginas do PDF
        for page in pdf_file:
            # percorre todas as linhas cdo PDF
            for line in page.get_text().splitlines():
                # verifica se a linha contém o número da matrícula
                if registration_number in line:
                    # encontra o retângulo que contém o número da matrícula
                    highlight = page.search_for(registration_number, hit_max=1)
                    if highlight:
                        highlight_rect = fitz.Rect(highlight[0][:4])

                        # destaca o número da matrícula no PDF
                        highlight_annot = page.add_highlight_annot(highlight_rect)

    # salva o PDF com os números de matrícula destacados
    output_filename = file_name
    file_number = 1
    while os.path.exists(output_filename):
        output_filename = f"{file_name.strip('.pdf')}({file_number}).pdf"
        file_number += 1

    pdf_file.save(output_filename)

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

root = ctk.CTk()
root.title("Destacar pdfs por matrícula")
# root.geometry("578x146")
root.resizable(False, False)

root.columnconfigure(0, weight=1)
root.rowconfigure(1, weight=1)

# Cria o frame para o arquivo Excel
excel_frame = ctk.CTkFrame(root)
excel_frame.grid(sticky=ctk.EW, padx=10, pady=10)

excel_file_label = ctk.CTkLabel(excel_frame, text="  Arquivo Excel:")
excel_file_label.grid(row=0, column=0, padx=5)

excel_file_entry = ctk.CTkEntry(excel_frame, placeholder_text="Selecione o arquivo Excel", width=300)
excel_file_entry.grid(row=0, column=1, padx=10)

excel_file_button = ctk.CTkButton(excel_frame, text="Selecionar", command=select_excel_file)
excel_file_button.grid(row=0, column=2, padx=5)

# Cria o frame para o arquivo PDF
pdf_frame = ctk.CTkFrame(root)
pdf_frame.grid(sticky=ctk.EW, padx=10, pady=10)

pdf_file_label = ctk.CTkLabel(pdf_frame, text="  Arquivo PDF:  ")
pdf_file_label.grid(row=0, column=0, padx=5)

pdf_file_entry = ctk.CTkEntry(pdf_frame, placeholder_text="Selecione o arquivo PDF", width=300)
pdf_file_entry.grid(row=0, column=1, padx=10)

pdf_file_button = ctk.CTkButton(pdf_frame, text="Selecionar", command=select_pdf_file)
pdf_file_button.grid(row=0, column=2, padx=5)

# Cria o botão para destacar os números de matrícula
highlight_button = ctk.CTkButton(root, text="Destacar", command=highlight_registration_numbers)
highlight_button.grid(row=2, column=0, sticky=ctk.EW, pady=5, padx=10)

# Frame para o painel de informações
info_frame = ctk.CTkFrame(root)
info_frame.columnconfigure(0, weight=1)

# Texto explicativo
info_text = """
Este programa permite que você destaque números de matrícula em um arquivo PDF
usando informações de uma planilha do Excel. Para isso, siga os passos abaixo:

1. Clique no botão 'Selecionar' ao lado de 'Arquivo Excel' para escolher o arquivo que
   contém os números de matrícula. Certifique-se de que os números estão na coluna B.

2. Clique no botão 'Selecionar' ao lado de 'Arquivo PDF' para escolher o arquivo PDF
   onde os números de matrícula serão destacados.

3. Depois de selecionar os arquivos, clique no botão 'Destacar' para iniciar o processo
   de destaque dos números de matrícula no PDF.

O PDF resultante com os números de matrícula destacados será salvo na mesma pasta do
arquivo PDF original. O nome do arquivo terá o mesmo nome do PDF original, seguido de
um número entre parênteses para evitar conflitos de nome em casos de processamento
repetido no mesmo PDF.
"""

info_label = ctk.CTkLabel(info_frame, text=info_text, wraplength=550, )
info_label.grid(row=0, column=0, padx=10, pady=10)

center_frame = ctk.CTkFrame(root)
center_frame.grid(row=4, column=0, columnspan=3, sticky=ctk.EW)

# Botão de expansão para mostrar/ocultar o painel de informações
info_expander = ctk.CTkButton(center_frame, text="Como funciona?", command=toggle_info_panel)
info_expander.pack(pady=5)

# Inicia a janela
root.mainloop()
