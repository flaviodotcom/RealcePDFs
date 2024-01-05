import os
import re

from openpyxl import load_workbook
from pypdf import PdfWriter, PdfReader

from realce.core.destacar import BaseRealcePdf
from realce.infra.logger import RealceLogger
from realce.infra.thread import WorkerThread


class SepararPDF(BaseRealcePdf):
    def __init__(self):
        self.logger = RealceLogger.get_logger()

    @staticmethod
    def separar_vt(campo_arquivo_excel, campo_arquivo_pdf, pasta_destino):
        nova_pasta_destino = SepararPDF.criar_pasta_vt_separado(pasta_destino)
        pdf, matriculas_nao_encontradas, nome_arquivo = SepararPDF.destacar_pdf(campo_arquivo_excel, campo_arquivo_pdf)

        nome_arquivo = SepararPDF.salvar_arquivo_pdf(nova_pasta_destino, nome_arquivo, pdf)
        pdf_reader = PdfReader(os.path.join(nova_pasta_destino, nome_arquivo))
        planilha = SepararPDF.separar_pdf_por_matricula(pdf_reader, campo_arquivo_excel, nova_pasta_destino)

        pdf_mesclado = SepararPDF.mesclar_pdfs(nova_pasta_destino, nome_arquivo, planilha)
        SepararPDF.salvar_arquivo_pdf(nova_pasta_destino, '- PDF_MERGEADO.pdf', pdf_mesclado)
        SepararPDF.exibir_mensagem_conclusao(nova_pasta_destino, matriculas_nao_encontradas)

    @staticmethod
    def criar_pasta_vt_separado(pasta_destino):
        pasta_destino = f'{pasta_destino}/Vt'
        check_folder = os.path.isdir(pasta_destino)
        if not check_folder:
            os.makedirs(pasta_destino)
        return pasta_destino

    @staticmethod
    def separar_pdf_por_matricula(pdf_reader, campo_arquivo_excel, pasta_destino):
        planilha = load_workbook(campo_arquivo_excel).active
        WorkerThread.currentThread().progressUpdated.emit(50)

        for linha in range(1, planilha.max_row + 1):
            numero_matricula = planilha.cell(row=linha, column=2).value
            nome_func = str(planilha.cell(row=linha, column=3).value)
            new_pdf = PdfWriter()

            if numero_matricula:
                matricula = str(numero_matricula)
                for pagina in pdf_reader.pages:
                    if matricula in pagina.extract_text():
                        new_pdf.add_page(pagina)

            SepararPDF.logger.info(f'Iterando... '
                                   f'Linha {linha}; Número da matrícula {numero_matricula}; Funcionário {nome_func}')
            if len(new_pdf.pages) > 0:
                output_file = f"{pasta_destino}/{nome_func.strip()}.pdf"
                with open(output_file, "wb") as f:
                    SepararPDF.logger.info(f'Salvando arquivo {f.name}')
                    new_pdf.write(f)

        return planilha

    @staticmethod
    def mesclar_pdfs(pasta_destino, arquivo_destacado, planilha):
        WorkerThread.currentThread().progressUpdated.emit(75)
        pdf_mergeado = PdfWriter()
        matriculas_adicionadas = set()

        with open(os.path.join(pasta_destino, arquivo_destacado), 'rb') as pdf:
            pdf_reader = PdfReader(pdf)

            for numero_pagina in range(len(pdf_reader.pages)):
                texto_pagina = re.sub(r'(_)(?=\d)', r'\1 ', pdf_reader.pages[numero_pagina].extract_text())

                procura_matriculas_excel = [str(planilha.cell(row=i, column=2).value) for i in
                                            range(1, planilha.max_row + 1)]

                matriculas_encontradas = set(procura_matriculas_excel) & set(texto_pagina.split())

                SepararPDF.logger.info(
                    f'Juntando pdfs... Matrículas encontradas na página {numero_pagina + 1}:'
                    f' {matriculas_encontradas if str(matriculas_encontradas) != "set()" else "Nenhuma"}')

                if matriculas_encontradas and not matriculas_encontradas.issubset(matriculas_adicionadas):
                    matriculas_adicionadas.update(matriculas_encontradas)
                    pdf_mergeado.add_page(pdf_reader.pages[numero_pagina])
                    SepararPDF.logger.info(f'Página adicionada')

        return pdf_mergeado
