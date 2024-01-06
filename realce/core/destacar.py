import os
import re

import fitz
import openpyxl
from pypdf import PdfWriter

from realce.infra.logger import get_logger
from realce.infra.thread import WorkerThread


class BaseRealcePdf:
    logger = get_logger('RealcePDFs')

    @staticmethod
    def destacar_pdf(campo_arquivo_excel, campo_arquivo_pdf):
        nome_arquivo = os.path.basename(campo_arquivo_pdf)

        arquivo_excel = openpyxl.load_workbook(campo_arquivo_excel)
        arquivo_pdf = fitz.open(campo_arquivo_pdf)
        pdf_text = [pagina.get_text() for pagina in arquivo_pdf.pages()]

        WorkerThread.currentThread().progressUpdated.emit(25)

        matriculas_nao_encontradas = list()
        BaseRealcePdf.logger.info('Começando o destaque de PDFs')

        for linha in range(1, arquivo_excel.active.max_row + 1):
            numero_matricula = arquivo_excel.active.cell(row=linha, column=2).value
            nome_funcionario = str(arquivo_excel.active.cell(row=linha, column=3).value).upper()

            encontrou_matricula = False

            if numero_matricula:
                matricula = str(numero_matricula)
                matricula_encontrada = any(matricula in texto_pagina for texto_pagina in pdf_text)
                if matricula_encontrada:
                    for pagina, texto_pagina in zip(arquivo_pdf.pages(), pdf_text):
                        if matricula in texto_pagina:
                            realce = pagina.search_for(matricula, hit_max=1)
                            if realce:
                                retangulo_realce = fitz.Rect(realce[0][:4])
                                pagina.add_highlight_annot(retangulo_realce)
                                encontrou_matricula = True
                                BaseRealcePdf.logger.info(
                                    f'Destacando a matrícula {matricula} do funcionário {nome_funcionario}')
                                break

                if nome_funcionario and not encontrou_matricula:
                    matriculas_nao_encontradas.append(f'{numero_matricula} - {nome_funcionario}')

        return arquivo_pdf, matriculas_nao_encontradas, nome_arquivo

    @staticmethod
    def salvar_arquivo_pdf(pasta_destino, nome_arquivo, arquivo_pdf):
        numero_arquivo = 1

        while os.path.exists(os.path.join(pasta_destino, nome_arquivo)):
            match = re.search(r'\((\d+)\)', nome_arquivo)

            if match:
                numero_arquivo_existente = int(match.group(1))
                novo_numero_arquivo = numero_arquivo_existente + 1
                nome_arquivo = re.sub(r'\((\d+)\)', f'({novo_numero_arquivo})', nome_arquivo)
            else:
                nome_arquivo = f"{os.path.splitext(nome_arquivo)[0]}({numero_arquivo}).pdf"

            numero_arquivo += 1

        BaseRealcePdf.logger.info(f'Salvando o arquivo {nome_arquivo}')
        caminho_arquivo_saida = os.path.join(pasta_destino, nome_arquivo)

        if isinstance(arquivo_pdf, PdfWriter):
            with open(caminho_arquivo_saida, "wb") as f:
                arquivo_pdf.write(f)
        else:
            WorkerThread.currentThread().progressUpdated.emit(50)
            arquivo_pdf.save(caminho_arquivo_saida)

        return nome_arquivo

    @staticmethod
    def exibir_mensagem_conclusao(pasta_destino, matriculas_nao_encontradas):
        BaseRealcePdf.logger.info(f'O Arquivo final foi salvo em: {pasta_destino}')
        WorkerThread.currentThread().progressUpdated.emit(99)

        if matriculas_nao_encontradas:
            nome_arquivo_txt = "Matrículas não encontradas.txt"
            numero_arquivo_txt = 1

            while os.path.exists(os.path.join(pasta_destino, nome_arquivo_txt)):
                nome_arquivo_txt = f"Matrículas não encontradas({numero_arquivo_txt}).txt"
                numero_arquivo_txt += 1

            caminho_arquivo_txt = os.path.join(pasta_destino, nome_arquivo_txt)

            with open(caminho_arquivo_txt, "w") as arquivo_txt:
                for matricula in matriculas_nao_encontradas:
                    arquivo_txt.write(matricula + "\n")

            BaseRealcePdf.logger.info('Algumas matrículas não foram encontradas')

        BaseRealcePdf.logger.info('Processo finalizado!')
        WorkerThread.currentThread().progressUpdated.emit(100)


class RealceMatriculas(BaseRealcePdf):

    @staticmethod
    def pdf(campo_arquivo_excel, campo_arquivo_pdf, pasta_destino):
        arquivo_pdf, matriculas_nao_encontradas, nome_arquivo = BaseRealcePdf.destacar_pdf(campo_arquivo_excel,
                                                                                           campo_arquivo_pdf)

        RealceMatriculas.salvar_arquivo_pdf(pasta_destino, nome_arquivo, arquivo_pdf)
        RealceMatriculas.exibir_mensagem_conclusao(pasta_destino, matriculas_nao_encontradas)
