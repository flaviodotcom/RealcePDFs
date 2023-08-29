import PyPDF2
import PyPDF4
import os
import fitz
import openpyxl
from tkinter import filedialog, messagebox
from alive_progress import alive_bar
import subprocess


def run_command_in_cmd(command):
    # Inicialize o processo do CMD
    process = subprocess.Popen(command, stdout=subprocess.PIPE, stderr=subprocess.PIPE, shell=True, universal_newlines=True)

    # Leia a saída do processo e imprima no CMD
    while True:
        output_line = process.stdout.readline()
        if not output_line:
            break
        print(output_line.strip())

    # Espere até que o processo termine
    process.wait()

def separar_vt(campo_arquivo_pdf, campo_arquivo_excel, nome_arquivo):
    pasta_destino = filedialog.askdirectory()
    if not pasta_destino:
        messagebox.showerror("Erro", "Por favor, selecione uma pasta de destino.")
        return

    caminho_arquivo_pdf = campo_arquivo_pdf.get()
    if not caminho_arquivo_pdf:
        messagebox.showerror("Erro", "Por favor, selecione o arquivo PDF.")
        return

    arquivo_pdf = fitz.open(caminho_arquivo_pdf)

    caminho_arquivo_excel = campo_arquivo_excel.get()
    if not caminho_arquivo_excel:
        messagebox.showerror("Erro", "Por favor, selecione o arquivo Excel.")
        return
    
    arquivo_excel = openpyxl.load_workbook(caminho_arquivo_excel)

    planilha = arquivo_excel.active

    num_linhas = planilha.max_row

    matriculas_nao_encontradas = []

    num_funcionarios = 0

    # Inicialize a barra de progresso fora do loop
    with alive_bar(num_linhas - 14, title='Destacando Matriculas...', bar='classic') as bar:
        for linha in range(9, num_linhas - 5):
            numero_matricula = planilha.cell(row=linha, column=2).value
            numero_matricula = str(numero_matricula)

            nome_matricula = planilha.cell(row=linha, column=3).value
            nome_matricula = str(nome_matricula).upper().split(" ", 3)
            nome_matricula = nome_matricula[0] + " " + nome_matricula[1] + " " + nome_matricula[2]

            encontrou_matricula = False

            # Atualize a barra de progresso a cada iteração de página
            bar()

            for pagina in arquivo_pdf:
                for linha_texto in pagina.get_text().splitlines():
                    if numero_matricula in linha_texto:
                        realce = pagina.search_for(numero_matricula, hit_max=1)
                        num_funcionarios += 1
                        if realce:
                            retangulo_realce = fitz.Rect(realce[0][:4])
                            pagina.add_highlight_annot(retangulo_realce)
                            encontrou_matricula = True
                            break

            if not encontrou_matricula and nome_matricula and numero_matricula != "None":
                matriculas_nao_encontradas.append(numero_matricula + " - " + nome_matricula)

    # Abrir o CMD e mostrar a barra de progresso lá
    cmd_command = 'echo Executando tarefa...'
    run_command_in_cmd(cmd_command)

    nome_arquivo_saida = nome_arquivo
    numero_arquivo = 1
    while os.path.exists(os.path.join(pasta_destino, nome_arquivo_saida)):
        nome_arquivo_saida = f"{nome_arquivo.strip('.pdf')}({numero_arquivo}).pdf"
        numero_arquivo += 1

    caminho_arquivo_saida = os.path.join(pasta_destino, nome_arquivo_saida)
    arquivo_pdf.save(caminho_arquivo_saida)

    arquivo_pdf = PyPDF2.PdfReader(caminho_arquivo_saida)

    pasta_destino = pasta_destino + "/Vt Separado"
    check_folder = os.path.isdir(pasta_destino)

    if not check_folder:
        os.makedirs(pasta_destino)

    with alive_bar(num_funcionarios, title='Separando PDFs...', bar='classic') as bar:
        for linha in range(1, num_linhas + 1):
            numero_matricula = planilha.cell(row=linha, column=2).value
            numero_matricula = str(numero_matricula)

            nome_func = planilha.cell(row=linha, column=3).value
            nome_func = str(nome_func)

            new_pdf = PyPDF2.PdfWriter()

            for pagina in arquivo_pdf.pages:
                if numero_matricula in pagina.extract_text():
                    new_pdf.add_page(pagina)

            if len(new_pdf.pages) > 0:
                output_file = f"{pasta_destino}/{nome_func}.pdf"
                with open(output_file, "wb") as f:
                    new_pdf.write(f)
                    bar()

    pasta = f"{pasta_destino}"

    pdf_mesclado = PyPDF4.PdfFileMerger()
    pdfs_encontrados = False
    for arquivo in os.listdir(pasta):
        if arquivo.lower().endswith(".pdf"):
            caminho_arquivo = os.path.join(pasta, arquivo)
            with open(caminho_arquivo, 'rb') as f:
                pdf_mesclado.append(f)
            pdfs_encontrados = True

    # Salvar arquivo mesclado com o nome da pasta selecionada
    if pdfs_encontrados:
        caminho_arquivo_mesclado = f"{pasta_destino}/- Vt final.pdf"
        with open(caminho_arquivo_mesclado, "wb") as saida:
            pdf_mesclado.write(saida)

        # Mensagem de conclusão
        messagebox.showinfo("Concluído", "PDFs mesclados e salvos em: " + caminho_arquivo_mesclado)

    if matriculas_nao_encontradas:
        nome_arquivo_txt = "Matrículas não encontradas.txt"
        numero_arquivo_txt = 1

        while os.path.exists(os.path.join(pasta_destino, nome_arquivo_txt)):
            nome_arquivo_txt = f"Matrículas não encontradas({numero_arquivo_txt}).txt"
            numero_arquivo_txt += 1

        caminho_arquivo_txt = os.path.join(pasta_destino, nome_arquivo_txt)

        with open(caminho_arquivo_txt, "w") as arquivo_txt:
            for matricula in matriculas_nao_encontradas:
                arquivo_txt.write(matricula + "\n")

    messagebox.showinfo("Concluído", f"O PDF editado foi salvo em:\n{pasta_destino}")

    messagebox.showwarning(
        "Matrículas não encontradas",
        f"Não foi possível encontrar algumas matrículas no arquivo PDF selecionado.\n\nFoi gerado um arquivo de texto que contém as matrículas não encontradas, salvo em:\n{caminho_arquivo_txt}",
    )

