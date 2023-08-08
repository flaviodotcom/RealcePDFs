import os
import fitz
import openpyxl
import customtkinter as ctk
from tkinter import filedialog, messagebox

nome_arquivo = ""

def selecionar_arquivo_excel():
    caminho_arquivo = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx"), ("Arquivos CSV", "*.csv"), ("Arquivos Excel 97-2003", "*.xls")])
    campo_arquivo_excel.delete(0, ctk.END)
    campo_arquivo_excel.insert(0, caminho_arquivo)

def selecionar_arquivo_pdf():
    caminho_arquivo = filedialog.askopenfilename(filetypes=[("Arquivos PDF", "*.pdf")])
    global nome_arquivo
    campo_arquivo_pdf.delete(0, ctk.END)
    campo_arquivo_pdf.insert(0, caminho_arquivo)
    nome_arquivo = os.path.basename(caminho_arquivo)

def alternar_painel_informacoes():
    if painel_informacoes.grid_info():
        painel_informacoes.grid_remove()
    else:
        painel_informacoes.grid(row=3, column=0, columnspan=3, padx=10, pady=10, sticky=ctk.EW)

def realcar_numeros_matricula(pasta_destino):
    global nome_arquivo
    caminho_arquivo_pdf = campo_arquivo_pdf.get()
    if not caminho_arquivo_pdf:
        ctk.messagebox.showerror("Erro", "Por favor, selecione o arquivo PDF.")
        return

    arquivo_pdf = fitz.open(caminho_arquivo_pdf)

    caminho_arquivo_excel = campo_arquivo_excel.get()
    if not caminho_arquivo_excel:
        ctk.messagebox.showerror("Erro", "Por favor, selecione o arquivo Excel.")
        return

    arquivo_excel = openpyxl.load_workbook(caminho_arquivo_excel)
    planilha = arquivo_excel.active

    num_linhas = planilha.max_row

    matriculas_nao_encontradas = []

    for linha in range(1, num_linhas + 1):
        numero_matricula = planilha.cell(row=linha, column=2).value
        numero_matricula = str(numero_matricula)

        nome_matricula = planilha.cell(row=linha, column=3).value
        nome_matricula = str(nome_matricula)

        encontrou_matricula = False

        for pagina in arquivo_pdf:
            for linha_texto in pagina.get_text().splitlines():
                if numero_matricula in linha_texto:
                    realce = pagina.search_for(numero_matricula, hit_max=1)
                    if realce:
                        retangulo_realce = fitz.Rect(realce[0][:4])
                        anotacao_realce = pagina.add_highlight_annot(retangulo_realce)
                        encontrou_matricula = True
                        break
        
        if not encontrou_matricula and nome_matricula and numero_matricula != 'None':
            matriculas_nao_encontradas.append(numero_matricula + " - " + nome_matricula)

    nome_arquivo_saida = nome_arquivo
    numero_arquivo = 1
    while os.path.exists(os.path.join(pasta_destino, nome_arquivo_saida)):
        nome_arquivo_saida = f"{nome_arquivo.strip('.pdf')}({numero_arquivo}).pdf"
        numero_arquivo += 1

    caminho_arquivo_saida = os.path.join(pasta_destino, nome_arquivo_saida)
    arquivo_pdf.save(caminho_arquivo_saida)

    if matriculas_nao_encontradas:
        nome_arquivo_txt = "Matrículas não encontradas.txt"
        numero_arquivo_txt = 1

        while os.path.exists(os.path.join(pasta_destino, nome_arquivo_txt)):
            nome_arquivo_txt = f"Matrículas não encontradas({numero_arquivo_txt}).txt"
            numero_arquivo_txt += 1

        caminho_arquivo_txt = os.path.join(pasta_destino, nome_arquivo_txt)
        
        with open(caminho_arquivo_txt, "w") as arquivo_txt:
            for matricula in matriculas_nao_encontradas:
                arquivo_txt.write(matricula + "\n")

    messagebox.showinfo("Concluído", f"O PDF editado foi salvo em:\n{caminho_arquivo_saida}")
    messagebox.showwarning("Matrículas não encontradas", f"Não foi possível encontrar algumas matrículas no arquivo PDF selecionado.\nFoi gerado um arquivo de texto que contém as matrículas não encontradas, salvo em:\n{caminho_arquivo_txt}")


def salvar_para_pasta_padrao():
    pasta_destino = os.path.join(os.path.dirname(__file__), "BENEFICIOS DESTACADOS")
    os.makedirs(pasta_destino, exist_ok=True)
    realcar_numeros_matricula(pasta_destino)

def salvar_para_pasta_selecionada_pelo_usuario():
    pasta_destino = filedialog.askdirectory()
    if not pasta_destino:
        return
    realcar_numeros_matricula(pasta_destino)

ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

root = ctk.CTk()
root.title("Destacar PDFs por matrícula")
root.resizable(False, False)

root.columnconfigure(0, weight=1)
root.rowconfigure(1, weight=1)

frame_excel = ctk.CTkFrame(root)
frame_excel.grid(sticky=ctk.EW, padx=10, pady=10)

rotulo_arquivo_excel = ctk.CTkLabel(frame_excel, text="  Arquivo Excel:")
rotulo_arquivo_excel.grid(row=0, column=0, padx=5)

campo_arquivo_excel = ctk.CTkEntry(frame_excel, placeholder_text="Selecione o arquivo Excel", width=300)
campo_arquivo_excel.grid(row=0, column=1, padx=10)

botao_selecionar_arquivo_excel = ctk.CTkButton(frame_excel, text="Selecionar", command=selecionar_arquivo_excel)
botao_selecionar_arquivo_excel.grid(row=0, column=2, padx=5)

frame_pdf = ctk.CTkFrame(root)
frame_pdf.grid(sticky=ctk.EW, padx=10, pady=10)

rotulo_arquivo_pdf = ctk.CTkLabel(frame_pdf, text="  Arquivo PDF:  ")
rotulo_arquivo_pdf.grid(row=0, column=0, padx=5)

campo_arquivo_pdf = ctk.CTkEntry(frame_pdf, placeholder_text="Selecione o arquivo PDF", width=300)
campo_arquivo_pdf.grid(row=0, column=1, padx=10)

botao_selecionar_arquivo_pdf = ctk.CTkButton(frame_pdf, text="Selecionar", command=selecionar_arquivo_pdf)
botao_selecionar_arquivo_pdf.grid(row=0, column=2, padx=5)

frame_salvar_e_info = ctk.CTkFrame(root)
frame_salvar_e_info.grid(row=2, column=0, columnspan=3, sticky=ctk.EW, pady=5, padx=10)

botao_destacar = ctk.CTkButton(frame_salvar_e_info, text="Salvar", command=salvar_para_pasta_padrao)
botao_destacar.grid(row=0, column=0, padx=5, sticky=ctk.EW)

botao_destacar_em_outra_pasta = ctk.CTkButton(frame_salvar_e_info, text="Salvar Como", command=salvar_para_pasta_selecionada_pelo_usuario)
botao_destacar_em_outra_pasta.grid(row=0, column=1, padx=5, sticky=ctk.EW)

frame_salvar_e_info.columnconfigure(0, weight=1)
frame_salvar_e_info.columnconfigure(1, weight=1)

painel_informacoes = ctk.CTkFrame(root)
painel_informacoes.columnconfigure(0, weight=1)

texto_informacoes = """
Este programa permite destacar a matrícula dos funcionários em um arquivo PDF usando informações de uma planilha do Excel. Principalmente usado para realçar os benefícios de Seguro de Vida, Plano Odontológico, Vale Transporte, Vale Alimentação e Vale Refeição.

Orientações:

1. Clique no botão 'Selecionar' para escolher o arquivo Excel e PDF, respectivamente.

2. Certifique-se de que as matrículas estejam na coluna B da planilha. O programa percorre pela segunda coluna (coluna B), garanta que nesse coluna não existam outras informações.

3. O programa cria um arquivo de texto, que aponta quais foram as matrículas não encontradas junto com o nome do colaborador. Para que essa funcionalidade ocorra como esperado, mantenha o nome dos colaboradores na terceira coluna (coluna C) da planilha. O arquivo de texto será salvo no mesmo diretório do PDF editado.

4. Depois de selecionar os arquivos, clique no botão 'Salvar' para salvar o PDF editado no mesmo diretório em que o programa está localizado, dentro de uma pasta que será criada, chamada 'BENEFICIOS DESTACADOS'. Ou clique no botão 'Salvar Como', para salvar o PDF editado no caminho que preferir.
"""

tamanho_da_fonte = 13
fonte_personalizada = ("Arial", tamanho_da_fonte)

rotulo_informacoes = ctk.CTkLabel(painel_informacoes, text=texto_informacoes, wraplength=520, justify="left", font=fonte_personalizada)
rotulo_informacoes.grid(row=0, column=0, padx=10, pady=5)
rotulo_informacoes.configure(text_color="white")

frame_central = ctk.CTkFrame(root)
frame_central.grid(row=4, column=0, columnspan=3, sticky=ctk.EW)
frame_central.configure(fg_color="transparent")

botao_expansor_informacoes = ctk.CTkButton(frame_central, text="Como funciona?", command=alternar_painel_informacoes)
botao_expansor_informacoes.pack(pady=5)

root.mainloop()
