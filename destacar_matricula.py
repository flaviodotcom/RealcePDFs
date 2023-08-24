import os
import fitz
import openpyxl
import sys
import customtkinter as ctk
from tkinter import filedialog, messagebox
from system.segunda_janela import SegundaJanela
import system.functions


nome_arquivo = ""


class ErroPdf(Exception):
    pass


class ErroExcel(Exception):
    pass


def resource_path(relative_path):
    base_path = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    return os.path.join(base_path, relative_path)


def selecionar_arquivo_excel():
    caminho_arquivo = filedialog.askopenfilename(
        filetypes=[
            ("Arquivos Excel", "*.xlsx"),
            ("Arquivos CSV", "*.csv"),
            ("Arquivos Excel 97-2003", "*.xls"),
        ]
    )
    campo_arquivo_excel.delete(0, ctk.END)
    campo_arquivo_excel.insert(0, caminho_arquivo)


def selecionar_arquivo_pdf():
    caminho_arquivo = filedialog.askopenfilename(filetypes=[("Arquivos PDF", "*.pdf")])
    global nome_arquivo
    campo_arquivo_pdf.delete(0, ctk.END)
    campo_arquivo_pdf.insert(0, caminho_arquivo)
    nome_arquivo = os.path.basename(caminho_arquivo)


def realcar_numeros_matricula(pasta_destino):
    global nome_arquivo, caminho_arquivo_txt

    caminho_arquivo_pdf = campo_arquivo_pdf.get()
    arquivo_pdf = fitz.open(caminho_arquivo_pdf)

    caminho_arquivo_excel = campo_arquivo_excel.get()
    arquivo_excel = openpyxl.load_workbook(caminho_arquivo_excel)

    planilha = arquivo_excel.active

    num_linhas = planilha.max_row

    matriculas_nao_encontradas = []

    for linha in range(1, num_linhas + 1):
        numero_matricula = planilha.cell(row=linha, column=2).value
        numero_matricula = str(numero_matricula)

        nome_matricula = planilha.cell(row=linha, column=3).value
        nome_matricula = str(nome_matricula).upper()

        encontrou_matricula = False

        for pagina in arquivo_pdf:
            for linha_texto in pagina.get_text().splitlines():
                if numero_matricula in linha_texto:
                    realce = pagina.search_for(numero_matricula, hit_max=1)
                    if realce:
                        retangulo_realce = fitz.Rect(realce[0][:4])
                        pagina.add_highlight_annot(retangulo_realce)
                        encontrou_matricula = True
                        break

        if not encontrou_matricula and nome_matricula and numero_matricula != "None":
            matriculas_nao_encontradas.append(numero_matricula + " - " + nome_matricula)

    nome_arquivo_saida = nome_arquivo
    numero_arquivo = 1
    while os.path.exists(os.path.join(pasta_destino, nome_arquivo_saida)):
        nome_arquivo_saida = f"{nome_arquivo.strip('.pdf')}({numero_arquivo}).pdf"
        numero_arquivo += 1

    caminho_arquivo_saida = os.path.join(pasta_destino, nome_arquivo_saida)
    arquivo_pdf.save(caminho_arquivo_saida)

    if matriculas_nao_encontradas:
        nome_arquivo_txt = "Matrículas não encontradas.txt"
        numero_arquivo_txt = 1

        while os.path.exists(os.path.join(pasta_destino, nome_arquivo_txt)):
            nome_arquivo_txt = f"Matrículas não encontradas({numero_arquivo_txt}).txt"
            numero_arquivo_txt += 1

        caminho_arquivo_txt = os.path.join(pasta_destino, nome_arquivo_txt)

        with open(caminho_arquivo_txt, "w") as arquivo_txt:
            for matricula in matriculas_nao_encontradas:
                arquivo_txt.write(matricula + "\n")

    messagebox.showinfo(
        "Concluído", f"O PDF editado foi salvo em:\n{caminho_arquivo_saida}"
    )

    messagebox.showwarning(
        "Matrículas não encontradas",
        f"Não foi possível encontrar algumas matrículas no arquivo PDF selecionado.\n\nFoi gerado um arquivo de texto que contém as matrículas não encontradas, salvo em:\n{caminho_arquivo_txt}",
    )


def tratar_erro(caminho_arquivo_excel, caminho_arquivo_pdf):
    try:
        mensagem_erro = (
            "Por favor, selecione o arquivo Excel e o arquivo PDF."
            if not caminho_arquivo_excel and not caminho_arquivo_pdf
            else "Por favor, selecione o arquivo PDF."
            if not caminho_arquivo_pdf
            else "Por favor, selecione o arquivo Excel."
            if not caminho_arquivo_excel
            else None
        )

        if mensagem_erro:
            messagebox.showerror("Erro", mensagem_erro)
            return True
        return False

    except Exception as e:
        messagebox.showerror("Erro", f"Ocorreu um erro: {str(e)}")
        return True


def tratar_erro_pdf(caminho_arquivo_pdf):
    try:
        fitz.open(caminho_arquivo_pdf)
    except Exception as e:
        raise ErroPdf(f"Erro no arquivo PDF: {str(e)}")


def tratar_erro_excel(caminho_arquivo_excel):
    try:
        openpyxl.load_workbook(caminho_arquivo_excel)
    except Exception as e:
        raise ErroExcel(f"Erro no arquivo Excel: {str(e)}")


def separar_vt():
    system.functions.separar_vt(campo_arquivo_pdf=campo_arquivo_pdf,
                                campo_arquivo_excel=campo_arquivo_excel,
                                nome_arquivo=nome_arquivo)

def salvar_para_pasta_padrao():
    caminho_arquivo_excel = campo_arquivo_excel.get()
    caminho_arquivo_pdf = campo_arquivo_pdf.get()

    try:
        tratar_erro(caminho_arquivo_excel, caminho_arquivo_pdf)
        pasta_destino = os.path.join(
            os.path.expanduser("~"), "Desktop", "BENEFICIOS DESTACADOS"
        )
        tratar_erro_pdf(caminho_arquivo_pdf)
        tratar_erro_excel(caminho_arquivo_excel)

        os.makedirs(pasta_destino, exist_ok=True)
        realcar_numeros_matricula(pasta_destino)
    except (ErroPdf, ErroExcel) as e:
        messagebox.showerror("Erro", str(e))


def salvar_para_pasta_selecionada_pelo_usuario():
    caminho_arquivo_excel = campo_arquivo_excel.get()
    caminho_arquivo_pdf = campo_arquivo_pdf.get()

    if tratar_erro(caminho_arquivo_excel, caminho_arquivo_pdf):
        return

    pasta_destino = filedialog.askdirectory()
    if not pasta_destino:
        return
    realcar_numeros_matricula(pasta_destino)


# Segunda Janela
sec_window = SegundaJanela()


def abrir_seg_janela():
    sec_window.abrir_janela()


def fechar():
    root.destroy()
    sec_window.fechar_seg_janela()


# Janela Principal
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("dark-blue")

root = ctk.CTk()
root.title(" Destacar PDFs por matrícula")
root.iconbitmap(resource_path("assets\\Cookie-Monster.ico"))
root.resizable(False, False)
root.protocol("WM_DELETE_WINDOW", fechar)

# Ajuste de tamanho e centralização
window_height = 220
window_width = 560

screen_height = root.winfo_screenheight()
screen_width = root.winfo_screenwidth()

x = (screen_width / 2) - (window_width / 2)
y = (screen_height / 2) - (window_height / 2)

root.geometry(f'{window_width}x{window_height}+{int(x)}+{int(y)}')

root.columnconfigure(0, weight=1)
root.rowconfigure(1, weight=1)

# Botões e Frames
frame_excel = ctk.CTkFrame(root)
frame_excel.grid(sticky=ctk.EW, padx=10, pady=10)

rotulo_arquivo_excel = ctk.CTkLabel(frame_excel, text="  Arquivo Excel:")
rotulo_arquivo_excel.grid(row=0, column=0, padx=3)

campo_arquivo_excel = ctk.CTkEntry(
    frame_excel, placeholder_text="Selecione o arquivo Excel", width=300
)
campo_arquivo_excel.grid(row=0, column=1, padx=5)

botao_selecionar_arquivo_excel = ctk.CTkButton(
    frame_excel, text="Selecionar", command=selecionar_arquivo_excel
)
botao_selecionar_arquivo_excel.grid(row=0, column=2)

frame_pdf = ctk.CTkFrame(root)
frame_pdf.grid(sticky=ctk.EW, padx=10, pady=10)

rotulo_arquivo_pdf = ctk.CTkLabel(frame_pdf, text="  Arquivo PDF:  ")
rotulo_arquivo_pdf.grid(row=0, column=0, padx=3)

campo_arquivo_pdf = ctk.CTkEntry(
    frame_pdf, placeholder_text="Selecione o arquivo PDF", width=300
)
campo_arquivo_pdf.grid(row=0, column=1, padx=5)

botao_selecionar_arquivo_pdf = ctk.CTkButton(
    frame_pdf, text="Selecionar", command=selecionar_arquivo_pdf
)
botao_selecionar_arquivo_pdf.grid(row=0, column=2)

frame_salvar_e_info = ctk.CTkFrame(root)
frame_salvar_e_info.grid(row=2, column=0, columnspan=3, sticky=ctk.EW, pady=5, padx=10)

botao_destacar = ctk.CTkButton(
    frame_salvar_e_info, text="Salvar", command=salvar_para_pasta_padrao
)
botao_destacar.grid(row=0, column=0, sticky=ctk.EW, padx=3)

botao_destacar_em_outra_pasta = ctk.CTkButton(
    frame_salvar_e_info,
    text="Salvar Como",
    command=salvar_para_pasta_selecionada_pelo_usuario,
)
botao_destacar_em_outra_pasta.grid(row=0, column=1, sticky=ctk.EW, padx=3)

frame_salvar_e_info.columnconfigure(0, weight=1)
frame_salvar_e_info.columnconfigure(1, weight=1)
frame_salvar_e_info.configure(fg_color="transparent")

separar_vts = ctk.CTkButton(frame_salvar_e_info, text="Separar PDFs por Matrícula", command=separar_vt)
separar_vts.grid(row=1, column=0, sticky=ctk.EW, padx=3, pady=10, columnspan=2)

botao_expansor_informacoes = ctk.CTkButton(frame_salvar_e_info, text="Como funciona?", command=abrir_seg_janela)
botao_expansor_informacoes.grid(row=2, column=0, columnspan=2, padx=3, pady=3)

root.mainloop()
