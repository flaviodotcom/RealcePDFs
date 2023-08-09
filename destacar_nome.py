import fitz
import openpyxl
import tkinter as tk
from tkinter import filedialog
from unidecode import unidecode

def select_excel_file():
    file_path = filedialog.askopenfilename(title="Selecione o arquivo Excel")
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(0, file_path)

def select_pdf_file():
    file_path = filedialog.askopenfilename(title="Selecione o arquivo PDF")
    pdf_file_entry.delete(0, tk.END)
    pdf_file_entry.insert(0, file_path)

def highlight_names():
    pdf_file = fitz.open(pdf_file_entry.get())

    excel_file = openpyxl.load_workbook(excel_file_entry.get())
    worksheet = excel_file.active

    num_rows = worksheet.max_row

    for row in range(1, num_rows + 1):
        name = worksheet.cell(row=row, column=1).value
        name = unidecode(name)

        if name not in [unidecode(text) for text in pdf_file.get_toc()]:
            for page in pdf_file:
                highlight = page.search_for(name, hit_max=1)
                if highlight:
                    highlight = highlight[0]
                    highlight_rect = fitz.Rect(highlight[:4])
                    highlight_annot = page.add_highlight_annot(highlight_rect)

    pdf_file.save('arquivoEditado.pdf')


root = tk.Tk()
root.title("Realçar nomes em PDF")

excel_frame = tk.Frame(root)
excel_frame.pack(fill=tk.X, padx=10, pady=5)

excel_file_label = tk.Label(excel_frame, text="Arquivo Excel:")
excel_file_label.pack(side=tk.LEFT)

excel_file_entry = tk.Entry(excel_frame, width=50)
excel_file_entry.pack(side=tk.LEFT, padx=5)

excel_file_button = tk.Button(excel_frame, text="Selecionar", command=select_excel_file)
excel_file_button.pack(side=tk.LEFT)

pdf_frame = tk.Frame(root)
pdf_frame.pack(fill=tk.X, padx=10, pady=5)

pdf_file_label = tk.Label(pdf_frame, text="Arquivo PDF:")
pdf_file_label.pack(side=tk.LEFT)

pdf_file_entry = tk.Entry(pdf_frame, width=50)
pdf_file_entry.pack(side=tk.LEFT, padx=5)

pdf_file_button = tk.Button(pdf_frame, text="Selecionar", command=select_pdf_file)
pdf_file_button.pack(side=tk.LEFT)

highlight_button = tk.Button(root, text="Realçar", command=highlight_names)
highlight_button.pack(pady=10)

root.mainloop()