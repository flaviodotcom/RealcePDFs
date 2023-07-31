import os
import fitz
import openpyxl
import tkinter as tk
from tkinter import filedialog

# Função para selecionar o arquivo Excel
def select_excel_file():
    file_path = filedialog.askopenfilename(title="Selecione o arquivo Excel")
    excel_file_entry.delete(0, tk.END)
    excel_file_entry.insert(0, file_path)


# Função para selecionar o arquivo PDF
def select_pdf_file():
    file_path = filedialog.askopenfilename(title="Selecione o arquivo PDF")
    pdf_file_entry.delete(0, tk.END)
    pdf_file_entry.insert(0, file_path)


# Função para destacar os números de matrícula no PDF
def highlight_registration_numbers():
    # abre o arquivo PDF
    pdf_file = fitz.open(pdf_file_entry.get())

    # abre a planilha do Excel
    excel_file = openpyxl.load_workbook(excel_file_entry.get())
    worksheet = excel_file.active

    # obtem o número de linhas na planilha
    num_rows = worksheet.max_row

    # percorre todas as linhas da planilha
    for row in range(1, num_rows + 1):
        # obtem o número da matrícula na coluna A da planilha
        registration_number = worksheet.cell(row=row, column=2).value
        registration_number = str(registration_number)  # converte para string

        # verifica se o número da matrícula não está presente no PDF
        if registration_number not in [text.lower() for text in pdf_file.get_toc()]:
            # destaca o número da matrícula no PDF
            for page in pdf_file:
                highlight = page.search_for(registration_number, hit_max=1)
                if highlight:
                    highlight = highlight[0]
                    highlight_rect = fitz.Rect(highlight[:4])
                    highlight_annot = page.add_highlight_annot(highlight_rect)

    # salva o PDF com os números de matrícula destacados
    output_filename = "realcePorMatricula.pdf"
    file_number = 1
    while os.path.exists(output_filename):
        output_filename = f"realcePorMatricula({file_number}).pdf"
        file_number += 1

    pdf_file.save(output_filename)


# Cria a janela principal
root = tk.Tk()
root.title("Realçar números de matrícula em PDF")

# Cria o frame para o arquivo Excel
excel_frame = tk.Frame(root)
excel_frame.pack(fill=tk.X, padx=10, pady=5)

excel_file_label = tk.Label(excel_frame, text="Arquivo Excel:")
excel_file_label.pack(side=tk.LEFT)

excel_file_entry = tk.Entry(excel_frame, width=50)
excel_file_entry.pack(side=tk.LEFT, padx=5)

excel_file_button = tk.Button(excel_frame, text="Selecionar", command=select_excel_file)
excel_file_button.pack(side=tk.LEFT)

# Cria o frame para o arquivo PDF
pdf_frame = tk.Frame(root)
pdf_frame.pack(fill=tk.X, padx=10, pady=5)

pdf_file_label = tk.Label(pdf_frame, text="Arquivo PDF:")
pdf_file_label.pack(side=tk.LEFT)

pdf_file_entry = tk.Entry(pdf_frame, width=50)
pdf_file_entry.pack(side=tk.LEFT, padx=5)

pdf_file_button = tk.Button(pdf_frame, text="Selecionar", command=select_pdf_file)
pdf_file_button.pack(side=tk.LEFT)

# Cria o botão para destacar os números de matrícula
highlight_button = tk.Button(root, text="Realçar", command=highlight_registration_numbers)
highlight_button.pack(pady=10)

# Inicia a janela
root.mainloop()
